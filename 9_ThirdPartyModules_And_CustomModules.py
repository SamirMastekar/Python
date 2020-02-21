#!/usr/bin/env python
# coding: utf-8

# ## Installing third party modules
# 



# py -m pip install --user <module_name>
# Run this from the command line to install the module for only the current user
# You can remove the --user argument to install it for all users if you have admin rights

# ## Web Scraping

# In[1]:


import requests
import webbrowser

# In[2]:


# this example HTML should mostly be present
import sys
example = sys.executable.replace('python.exe', r'Lib\idlelib\help.html')
# open any locally created HTML file in your browser, if it exists
default_brows = webbrowser.get()
default_brows.open(example)

# In[3]:


# to make a request
response = requests.get('https://www.reddit.com/r/Python/')

# check the response headers and the final URL!
print(response.reason, response.url)


# In[4]:


import requests
import bs4 # beautiful soup
import webbrowser

link = 'http://pntaconline/'
response = requests.get(link)
soup = bs4.BeautifulSoup(response.text)

a_tags = soup.select('pre')
for x in a_tags:
    print(x.text)

# # Find costs from Amazon India website
link = 'https://www.amazon.in/'
response = requests.get(link)
soup = bs4.BeautifulSoup(response.text)

deals = soup.find_all('span', class_="a-list-item")
for deal in deals:
    name = deal.find('img').get('alt')
    price = deal.find('div', class_='dealPrice')
    if price:
        price = price.text
        print(name, price)
        webbrowser.open(deal.find('img').get('src'))

# ## Read and Write Excel



# Read and Write a workbook
from openpyxl import load_workbook

mysheet = load_workbook(r"ancilliaryFiles/input.xlsx")

wb = mysheet.get_active_sheet()
for one, two, three in wb['A1': 'C5']:
#     print(dir(a))
    print(one.value, two.value, three.value)
    three.value = f'*******{three.value.strip()}*******'
mysheet.save('output.xlsx')



from openpyxl import Workbook



import os
import pprint
pprint.pprint(list(os.environ.values()))



# Dump all the environment variables in an excel sheet
import os
from openpyxl import Workbook
wb = Workbook()
default_sheet = wb.active
for key, value in os.environ.items():
    default_sheet.append((key, value))
wb.save('environ_dump.xlsx')

# ## Custom Modules



# If your module is not in the cwd then you can temporarily add that
# path to your PATH variable
import sys
sys.path.append('ancilliaryFiles')
import my_module



# to plot graphs, you can use matplotlib
from matplotlib import pyplot as plt



x = range(10)
y = range(10)



plt.plot(x, y, 'bo-')
plt.xlabel("X value")
plt.ylabel("Y value")
plt.title("y = x")
plt.savefig("test.png")



import psutil
import time
# by default it will consider the current Process (python)
edge = psutil.Process()
x = list()
y = list()
counter = 0

while counter < 10:
    x.append(counter)
    y.append(edge.cpu_times().user)
    plt.plot(x, y, 'bo-')
    plt.xlabel("Seconds")
    plt.ylabel("User value of CPU")
    plt.title("Monitoring a process")
    plt.show()
    time.sleep(1)
    counter += 1

# In[5]:







